from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.errors import ConflictError, DomainError, NotFoundError
from app.models.enums import AuditEventType, ManifestStatus, ShipmentStatus
from app.models.manifest import Manifest, ManifestFile, ManifestRow, ManifestValidationError
from app.models.sample import SampleAlias, SampleIdentifier
from app.models.shipment import Shipment, ShipmentSample
from app.services.audit_service import audit_service
from app.services.sample_service import sample_service
from app.storage.local import storage

MAPPING_PATH = Path(__file__).resolve().parent.parent / "manifests" / "canonical_mapping.json"


def _load_mapping_config() -> dict[str, Any]:
    return json.loads(MAPPING_PATH.read_text())


def _detect_file_type(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    raise DomainError("UNSUPPORTED_FILE_TYPE", "Only CSV and XLSX manifests are supported.")


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = [{k: (v or "").strip() for k, v in row.items() if k} for row in reader]
    return headers, rows


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    parsed: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = raw[i] if i < len(raw) else None
            if value is None:
                item[header] = ""
            elif isinstance(value, datetime):
                item[header] = value.date().isoformat()
            elif isinstance(value, date):
                item[header] = value.isoformat()
            else:
                item[header] = str(value).strip()
        parsed.append(item)
    return headers, parsed


def _build_column_mapping(headers: list[str], config: dict[str, Any]) -> dict[str, str | None]:
    normalized = {h.strip().lower().replace(" ", "_"): h for h in headers}
    mapping: dict[str, str | None] = {}
    for canonical, aliases in config["aliases"].items():
        found = None
        for alias in aliases:
            key = alias.lower()
            if key in normalized:
                found = normalized[key]
                break
        mapping[canonical] = found
    return mapping


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


class ManifestService:
    def upload(
        self,
        db: Session,
        *,
        actor_id: UUID,
        filename: str,
        content_type: str | None,
        data: bytes,
    ) -> Manifest:
        file_type = _detect_file_type(filename)
        checksum = _sha256(data)
        config = _load_mapping_config()
        storage_key = f"manifests/{uuid4()}/{filename}"
        storage.save(storage_key, data)

        manifest = Manifest(
            original_filename=filename,
            file_type=file_type,
            size_bytes=len(data),
            checksum_sha256=checksum,
            status=ManifestStatus.PARSING.value,
            storage_key=storage_key,
            created_by=actor_id,
            uploaded_at=datetime.now(timezone.utc),
        )
        db.add(manifest)
        db.flush()
        db.add(
            ManifestFile(
                manifest_id=manifest.id,
                storage_key=storage_key,
                filename=filename,
                content_type=content_type,
            )
        )

        try:
            headers, source_rows = _parse_csv(data) if file_type == "csv" else _parse_xlsx(data)
        except Exception as exc:
            manifest.status = ManifestStatus.VALIDATION_FAILED.value
            db.add(
                ManifestValidationError(
                    manifest_id=manifest.id,
                    code="PARSE_ERROR",
                    message=f"Failed to parse file: {exc}",
                )
            )
            db.commit()
            raise DomainError("MANIFEST_PARSE_ERROR", f"Failed to parse manifest: {exc}") from exc

        mapping = _build_column_mapping(headers, config)
        manifest.column_mapping = mapping
        manifest.row_count = len(source_rows)
        for idx, source in enumerate(source_rows, start=1):
            canonical = {}
            for field in config["canonical_fields"]:
                source_col = mapping.get(field)
                canonical[field] = source.get(source_col, "").strip() if source_col else ""
            db.add(
                ManifestRow(
                    manifest_id=manifest.id,
                    row_number=idx,
                    source_data=source,
                    canonical_data=canonical,
                )
            )
        manifest.status = ManifestStatus.UPLOADED.value
        audit_service.record(
            db,
            event_type=AuditEventType.MANIFEST_UPLOADED,
            entity_type="Manifest",
            entity_id=manifest.id,
            actor_user_id=actor_id,
            after_state={
                "filename": filename,
                "checksum": checksum,
                "row_count": manifest.row_count,
                "mapping": mapping,
            },
        )
        db.commit()
        db.refresh(manifest)
        return self.get(db, manifest.id)

    def get(self, db: Session, manifest_id: UUID) -> Manifest:
        manifest = db.scalar(
            select(Manifest)
            .options(
                selectinload(Manifest.files),
                selectinload(Manifest.rows).selectinload(ManifestRow.errors),
                selectinload(Manifest.validation_errors),
            )
            .where(Manifest.id == manifest_id)
        )
        if manifest is None:
            raise NotFoundError("MANIFEST_NOT_FOUND", "Manifest not found")
        return manifest

    def list(self, db: Session) -> list[Manifest]:
        return list(db.scalars(select(Manifest).order_by(Manifest.uploaded_at.desc())))

    def validate(self, db: Session, manifest_id: UUID, actor_id: UUID) -> Manifest:
        manifest = self.get(db, manifest_id)
        if manifest.status == ManifestStatus.COMMITTED:
            raise ConflictError("MANIFEST_ALREADY_COMMITTED", "Committed manifests cannot be re-validated.")
        config = _load_mapping_config()
        for err in list(manifest.validation_errors):
            db.delete(err)
        db.flush()

        seen_sample_ids: dict[str, int] = {}
        seen_barcodes: dict[str, int] = {}
        id_pattern = re.compile(config["sample_id_pattern"])
        barcode_pattern = re.compile(config["barcode_pattern"])
        allowed_units = {u.lower() for u in config["allowed_units"]}
        valid_count = 0
        invalid_count = 0

        for row in sorted(manifest.rows, key=lambda r: r.row_number):
            errors: list[ManifestValidationError] = []
            data = row.canonical_data or {}

            def add(code: str, message: str, field: str | None = None) -> None:
                errors.append(
                    ManifestValidationError(
                        manifest_id=manifest.id,
                        row_id=row.id,
                        row_number=row.row_number,
                        field_name=field,
                        code=code,
                        message=message,
                    )
                )

            required = [
                "shipment_reference",
                "sample_external_id",
                "external_barcode",
                "sample_type",
                "quantity",
                "quantity_unit",
            ]
            for field in required:
                if not str(data.get(field) or "").strip():
                    add("REQUIRED_FIELD", f"{field} is required.", field)

            sample_ext = (data.get("sample_external_id") or "").strip()
            barcode = (data.get("external_barcode") or "").strip()
            if sample_ext and not id_pattern.match(sample_ext):
                add("INVALID_SAMPLE_ID_FORMAT", "Sample ID format is invalid.", "sample_external_id")
            if barcode and not barcode_pattern.match(barcode):
                add("INVALID_BARCODE_FORMAT", "Barcode format is invalid.", "external_barcode")

            qty_raw = (data.get("quantity") or "").strip()
            if qty_raw:
                try:
                    qty = Decimal(qty_raw)
                    if qty <= 0:
                        add("QUANTITY_NOT_POSITIVE", "Quantity must be greater than zero.", "quantity")
                except InvalidOperation:
                    add("QUANTITY_NOT_NUMERIC", "Quantity must be numeric.", "quantity")

            unit = (data.get("quantity_unit") or "").strip()
            if unit and unit.lower() not in allowed_units and unit not in config["allowed_units"]:
                add("INVALID_UNIT", f"Unit '{unit}' is not allowed.", "quantity_unit")

            collected = _parse_date(data.get("collection_date") or "")
            received = _parse_date(data.get("received_date") or "")
            if (data.get("collection_date") or "").strip() and collected is None:
                add("INVALID_DATE", "collection_date is not a valid date.", "collection_date")
            if (data.get("received_date") or "").strip() and received is None:
                add("INVALID_DATE", "received_date is not a valid date.", "received_date")
            if collected and received and received < collected:
                add("INVALID_RECEIVED_DATE", "received_date cannot be before collection_date.", "received_date")

            if sample_ext:
                if sample_ext in seen_sample_ids:
                    add(
                        "DUPLICATE_SAMPLE_ID_IN_UPLOAD",
                        f"Duplicate sample ID in upload (also row {seen_sample_ids[sample_ext]}).",
                        "sample_external_id",
                    )
                else:
                    seen_sample_ids[sample_ext] = row.row_number
                existing_ext = db.scalar(
                    select(SampleIdentifier).where(
                        SampleIdentifier.identifier_type == "EXTERNAL",
                        SampleIdentifier.value == sample_ext,
                    )
                )
                if existing_ext:
                    add("DUPLICATE_EXTERNAL_ID_IN_DB", "External sample ID already exists.", "sample_external_id")

            if barcode:
                if barcode in seen_barcodes:
                    add(
                        "DUPLICATE_BARCODE_IN_UPLOAD",
                        f"Duplicate barcode in upload (also row {seen_barcodes[barcode]}).",
                        "external_barcode",
                    )
                else:
                    seen_barcodes[barcode] = row.row_number
                existing_bc = db.scalar(
                    select(SampleAlias).where(SampleAlias.alias_type == "BARCODE", SampleAlias.value == barcode)
                )
                if existing_bc:
                    add("DUPLICATE_BARCODE_IN_DB", "Barcode already exists in the database.", "external_barcode")

            row.is_valid = len(errors) == 0
            for err in errors:
                db.add(err)
            if row.is_valid:
                valid_count += 1
            else:
                invalid_count += 1

        manifest.valid_row_count = valid_count
        manifest.invalid_row_count = invalid_count
        manifest.validated_at = datetime.now(timezone.utc)
        manifest.status = (
            ManifestStatus.VALIDATED.value if invalid_count == 0 else ManifestStatus.VALIDATION_FAILED.value
        )
        audit_service.record(
            db,
            event_type=AuditEventType.MANIFEST_VALIDATED,
            entity_type="Manifest",
            entity_id=manifest.id,
            actor_user_id=actor_id,
            after_state={
                "status": manifest.status,
                "valid_row_count": valid_count,
                "invalid_row_count": invalid_count,
            },
        )
        db.commit()
        db.expire_all()
        return self.get(db, manifest.id)

    def commit(self, db: Session, manifest_id: UUID, actor_id: UUID) -> Manifest:
        manifest = db.scalar(select(Manifest).where(Manifest.id == manifest_id).with_for_update())
        if manifest is None:
            raise NotFoundError("MANIFEST_NOT_FOUND", "Manifest not found")
        if manifest.status == ManifestStatus.COMMITTED:
            raise ConflictError("MANIFEST_ALREADY_COMMITTED", "Manifest has already been committed.")
        if manifest.status not in {ManifestStatus.VALIDATED, ManifestStatus.VALIDATION_FAILED}:
            raise DomainError("MANIFEST_NOT_VALIDATED", "Validate the manifest before commit.")

        rows = list(
            db.scalars(
                select(ManifestRow)
                .where(ManifestRow.manifest_id == manifest.id)
                .order_by(ManifestRow.row_number)
            )
        )
        valid_rows = [r for r in rows if r.is_valid]
        if not valid_rows:
            raise DomainError("NO_VALID_ROWS", "There are no valid rows to commit.")

        try:
            first = valid_rows[0].canonical_data or {}
            shipment_ref = first.get("shipment_reference") or f"SHP-{manifest.id.hex[:8].upper()}"
            existing = db.scalar(select(Shipment).where(Shipment.shipment_reference == shipment_ref))
            if existing:
                shipment = existing
            else:
                shipment = Shipment(
                    shipment_reference=shipment_ref,
                    status=ShipmentStatus.RECEIVED.value,
                    source_location=first.get("source_location") or None,
                    temperature_requirement=first.get("temperature_requirement") or None,
                    manifest_id=manifest.id,
                    created_by=actor_id,
                )
                db.add(shipment)
                db.flush()
                audit_service.record(
                    db,
                    event_type=AuditEventType.SHIPMENT_CREATED,
                    entity_type="Shipment",
                    entity_id=shipment.id,
                    actor_user_id=actor_id,
                    after_state={"shipment_reference": shipment.shipment_reference},
                )

            for row in valid_rows:
                data = row.canonical_data or {}
                qty = Decimal(str(data["quantity"]))
                unit = data["quantity_unit"]
                if unit.lower() == "ul":
                    unit = "uL"
                elif unit.lower() == "ml":
                    unit = "mL"
                sample = sample_service.create_sample(
                    db,
                    actor_id=actor_id,
                    sample_type=data["sample_type"],
                    quantity=qty,
                    quantity_unit=unit,
                    material_type=data.get("material_type") or None,
                    external_id=data.get("sample_external_id") or None,
                    barcode=data.get("external_barcode") or None,
                    collection_date=_parse_date(data.get("collection_date") or ""),
                    received_date=_parse_date(data.get("received_date") or ""),
                    source_location=data.get("source_location") or None,
                    temperature_requirement=data.get("temperature_requirement") or None,
                    shipment_id=shipment.id,
                )
                db.add(ShipmentSample(shipment_id=shipment.id, sample_id=sample.id))
                row.committed_sample_id = sample.id

            manifest.status = ManifestStatus.COMMITTED.value
            manifest.committed_at = datetime.now(timezone.utc)
            manifest.committed_shipment_id = shipment.id
            audit_service.record(
                db,
                event_type=AuditEventType.MANIFEST_COMMITTED,
                entity_type="Manifest",
                entity_id=manifest.id,
                actor_user_id=actor_id,
                after_state={
                    "shipment_id": str(shipment.id),
                    "committed_samples": len(valid_rows),
                },
            )
            db.commit()
        except Exception:
            db.rollback()
            raise
        db.expire_all()
        return self.get(db, manifest.id)

    def validation_csv(self, db: Session, manifest_id: UUID) -> str:
        manifest = self.get(db, manifest_id)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["row_number", "field_name", "code", "message"])
        for err in sorted(manifest.validation_errors, key=lambda e: (e.row_number or 0, e.field_name or "")):
            writer.writerow([err.row_number, err.field_name, err.code, err.message])
        return output.getvalue()


manifest_service = ManifestService()
